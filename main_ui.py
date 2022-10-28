import os, sys, time
from datetime import datetime
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *

import logging
from colorama import Fore, Back, Style

import pyvisa as visa
from openpyxl import load_workbook, Workbook

# UI/UX 
from untitled import *

# import from lib 
from lib.logcolors import bcolors
from lib.Thread_lib import Runthread
from lib.Thread_FFT_lib import Runthread_FFT
from lib.Autoreport import Autoreport_Runthread
from lib.Manual_thread_lib import Manual_Runthread
from lib.log_lib import *
from lib.upload_too import Upload_Runthread

class mainProgram(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(mainProgram, self).__init__(parent)
        self.setupUi(self)
        self.UI_default_setup()

        self.timestamp = None
        self.wb = None
        self.continue_single = 0

        self.temperature = ["-20", "25", "50"]
        self.temperature_index_list = {"-20":0, "25":1, "50":2}
        self.temperature_index = 0
        self.test_item = ["Regulation", "Load","Line","Eff"]
        self.switch_index = 0
        self.config = {}

        self.set_comboBox_sheet(self.comboBox_temp, self.temperature)

        # Push Button
        # VISA ADDRESS REFRESH FUNCTION
        self.Refresh_VISA(self.comboBox_TK_visa)
        self.Refresh_VISA(self.comboBox_PS_visa)
        self.Refresh_VISA(self.comboBox_EL_visa)

        self.pushButton_TK_visa.clicked.connect(self.setup_TK_addres)
        self.pushButton_PS_visa.clicked.connect(self.setup_PS_addres)
        self.pushButton_EL_visa.clicked.connect(self.setup_EL_addres)

        self.pushButton_Oc_label.clicked.connect(self.refresh_info)
        self.pushButton_ch_read.clicked.connect(self.read_label)
        self.pushButton_ch_write.clicked.connect(self.write_label)

        self.pushButton_testplan.clicked.connect(self.open_testplan)
        self.pushButton_template.clicked.connect(self.open_template)

        self.pushButton_open_upload.clicked.connect(self.open_uploadfile)

        self.pushButton_RUN.clicked.connect(self.btn_run)
        self.pushButton_Cancel.clicked.connect(self.thread_stop)
        #self.pushButton_Cancel.clicked.connect(self.close)

        self.pushButton_FF_Autotesting.clicked.connect(self.FFT_autotesting)

        self.excel_data = []

        self.set_logger()
        self.statusBar.showMessage('Started successfully.')
    
    def UI_default_setup(self):
        #load settings
        self.settings = QtCore.QSettings("config.ini", QtCore.QSettings.IniFormat)
        self.lineEdit_time_scale.setText(self.settings.value("DEFaultsetup/Time_Scale"))
        self.DSB_V_Scale_CH1.setValue(float(self.settings.value("DEFaultsetup/Voltage_Scale_CH1")))
        self.DSB_V_Scale_CH2.setValue(float(self.settings.value("DEFaultsetup/Voltage_Scale_CH2")))
        self.DSB_V_Scale_CH3.setValue(float(self.settings.value("DEFaultsetup/Voltage_Scale_CH3")))
        self.DSB_offset_CH1.setValue(float(self.settings.value("DEFaultsetup/Offset_CH1")))
        self.DSB_offset_CH2.setValue(float(self.settings.value("DEFaultsetup/Offset_CH2")))
        self.DSB_offset_CH3.setValue(float(self.settings.value("DEFaultsetup/Offset_CH3")))
        self.DSB_postion_CH1.setValue(float(self.settings.value("DEFaultsetup/Position_CH1")))
        self.DSB_postion_CH2.setValue(float(self.settings.value("DEFaultsetup/Position_CH2")))
        self.DSB_postion_CH3.setValue(float(self.settings.value("DEFaultsetup/Position_CH3")))
        self.SB_display_wavform.setValue(float(self.settings.value("DEFaultsetup/Display_Wavform")))
        self.SB_display_graticule.setValue(float(self.settings.value("DEFaultsetup/Display_Graticule")))

    def set_config(self):
        self.config = {
            "Time Scale": self.lineEdit_time_scale.text(),
            "Display wavfrom": self.SB_display_wavform.value(),
            "Display graticule": self.SB_display_graticule.value(),
            "Channel 1":{
                    "Voltage Scale":self.DSB_V_Scale_CH1.value(),
                    "OFFSet":self.DSB_offset_CH1.value(),
                    "Position":self.DSB_postion_CH1.value()
                },
            "Channel 2":{
                    "Voltage Scale":self.DSB_V_Scale_CH2.value(),
                    "OFFSet":self.DSB_offset_CH2.value(),
                    "Position":self.DSB_postion_CH2.value()
                },
            "Channel 3":{
                    "Voltage Scale":self.DSB_V_Scale_CH3.value(),
                    "OFFSet":self.DSB_offset_CH3.value(),
                    "Position":self.DSB_postion_CH3.value()
                }
        }

    def set_logger(self):
        now = datetime.now().strftime('%Y%m%d%H%M%S')
        logging.info("Setup Log File to "+ setup_loggers(now, prefix="./logs"))
        add_text_handler(self.logger_callback)

    def logger_callback(self, msg):
        try:
            self.textBrowser_log.insertPlainText(msg+"\n")
            #self.textBrowser_log.append(msg)
            self.textBrowser_log.verticalScrollBar().setValue(
            self.textBrowser_log.verticalScrollBar().maximum())
            self.textBrowser_log.update()
            #self.textBrowser_log.moveCursor(QtGui.QTextCursor.End)
        except Exception:
            pass

    def write_label(self):
        try:
            self.thread = Manual_Runthread()
            self.thread.switch = "3"
            self.thread.TK_VISA_ADDRESS = self.comboBox_TK_visa.currentText()
            self.thread.label_tmp = [(f'"{self.lineEdit_ch1.text()}"'),
                                    (f'"{self.lineEdit_ch2.text()}"'),
                                    (f'"{self.lineEdit_ch3.text()}"'),
                                    (f'"{self.lineEdit_ch4.text()}"'),]
            self.thread.start()
        except Exception:
            pass

    def read_label(self):
        try:
            self.thread = Manual_Runthread()
            self.thread.switch = "2"
            self.thread.TK_VISA_ADDRESS = self.comboBox_TK_visa.currentText()
            self.thread._CH_label.connect(self.ch_label)
            self.thread.start()
        except Exception:
            pass

    def ch_label(self, msg):
        print(msg)
        self.lineEdit_ch1.setText(msg[0].strip('"'))
        self.lineEdit_ch2.setText(msg[1].strip('"'))
        self.lineEdit_ch3.setText(msg[2].strip('"'))
        self.lineEdit_ch4.setText(msg[3].strip('"'))
    
    def refresh_info(self):
        self.thread = Manual_Runthread()
        self.thread.TK_VISA_ADDRESS = self.comboBox_TK_visa.currentText()
        self.thread.PW_VISA_ADDRESS = self.comboBox_PS_visa.currentText()
        self.thread.LD_VISA_ADDRESS = self.comboBox_EL_visa.currentText()
        self.thread.switch = "1"
        self.thread._device_info.connect(self.set_device_info)
        self.thread.start()

    def check_folder(self):
        if not os.path.isfile("./Measurement data/"+self.timestamp):
            os.mkdir( "./Measurement data/"+self.timestamp)
    
    def setup_timestamp(self):
        nowTime = int(time.time())
        struct_time = time.localtime(nowTime)
        self.timestamp = str(time.strftime("%Y%m%d%H%M%S", struct_time))

    def btn_run(self):
        self.temperature_index = self.temperature_index_list[self.comboBox_temp.currentText()]
        if self.temperature_index == 0:
            self.setup_timestamp()
            self.check_folder()
        self.switch_index = 0
        self.statusBar.showMessage('Load excel : {}'.format(self.test_item[0]))
        self.load_excel(self.test_item[0])
        self.autotest()
    
    def autotest(self):
        #self.set_config()
        self.thread = Runthread()
        self.thread.testype = self.test_item[self.switch_index]
        self.thread.TK_VISA_ADDRESS = self.comboBox_TK_visa.currentText()
        self.thread.PW_VISA_ADDRESS = self.comboBox_PS_visa.currentText()
        self.thread.LD_VISA_ADDRESS = self.comboBox_EL_visa.currentText()
        self.thread.time_scale = self.lineEdit_time_scale.text()
        self.thread.display_wavform = self.SB_display_wavform.value()
        self.thread.display_graticule = self.SB_display_graticule.value()
        self.thread.lineEdit_scale_period = self.lineEdit_scale_period.text()
        self.thread.Default_scale = self.lineEdit_FFT_scale.text()
        #self.thread.config = self.config
        self.thread.excel_data = self.excel_data
        self.thread.timestamp = self.timestamp
        self.thread.temperature_index = self.temperature_index
        self.thread._respones.connect(self.respones2table)
        self.thread._device_info.connect(self.set_device_info)
        self.thread._stop_signal.connect(self.switch_table)
        self.thread.start()
        self.statusBar.showMessage('{} Thread Runing ...'.format(self.test_item[self.switch_index]))
    
    def FFT_autotesting(self):
        self.label_freq_value.setText("")
        self.label_dBV_value.setText("")

        self.thread = Runthread_FFT()
        self.thread.lineEdit_scale_period = self.lineEdit_scale_period.text()
        self.thread.Default_scale = self.lineEdit_FFT_scale.text()
        self.thread._respones.connect(self.FFT_respones)
        self.thread.start()
    
    def FFT_respones(self, msg):
        self.label_freq_value.setText("%s Hz" %(msg[0]))
        self.label_dBV_value.setText("%s dBV" %(msg[1]))

    def set_device_info(self, msg):
        self.label_Oc_name_2.setText(msg[0])
        self.label_PS_name_2.setText(msg[1])
        self.label_EL_name_2.setText(msg[2])
    
    def table2excel(self):
        if self.switch_index == 0 and self.temperature_index == 0:
            wb_data = load_workbook(self.lineEdit_testplan.text(), data_only=True)
            basicsheet = wb_data.copy_worksheet(wb_data["Basic"])
            basicsheet.title = "Testing"
        elif self.switch_index > 0 or self.temperature_index <= 3:
            wb_data = load_workbook("Measurement data/"+self.timestamp+"/testingdata_"+self.timestamp+".xlsx", data_only=True)
            basicsheet = wb_data["Testing"]

        index_dic = {"Regulation": 4 , "Load": 24, "Line":40, "Eff":57}
        sheet_con = 0
        type_name = self.test_item[self.switch_index]
        #print(index_dic[type_name] , type_name)
        col = self.tableWidget_testplan.columnCount()
        row = self.tableWidget_testplan.rowCount()
        #print(col,row)
        for row_index in range(row):
            if row_index > 0:
                sheet_con += 1

                #print(sheet_con)
                logging.debug(str(self.tableWidget_testplan.item(row_index, 0).text()))
                for col_index in range(col):
                    teext = str(self.tableWidget_testplan.item(row_index, col_index).text())
                    basicsheet.cell(row=((row-1)*self.temperature_index)+(sheet_con+3), column=col_index+index_dic[type_name]).value = teext
        basicsheet.cell(row=1, column=3).value = self.label_Oc_name_2.text()
        basicsheet.cell(row=2, column=3).value = self.label_PS_name_2.text()
        basicsheet.cell(row=3, column=3).value = self.label_EL_name_2.text()
        wb_data.save("Measurement data/"+self.timestamp+"/testingdata_"+self.timestamp+".xlsx")

    def rum_autoreport(self):
        try: 
            self.thread_autoreport = Autoreport_Runthread()
            self.thread_autoreport.timestamp = self.timestamp
            self.thread_autoreport.Templatepath = self.lineEdit_template.text()
            self.thread_autoreport.testplan_path = self.lineEdit_testplan.text()
            self.thread_autoreport.start()
            self.statusBar.showMessage('Auto Report Thread Runing ...')
        except Exception as e:
            logging.error(e)

    def switch_table(self):
        self.table2excel()
        self.switch_index += 1
        if self.switch_index < len(self.test_item):
            self.load_excel(self.test_item[self.switch_index])
            self.autotest()
        if self.switch_index == len(self.test_item) and self.temperature_index == 2:
            self.rum_autoreport()
            logging.info("Done !")

    def w2table(self):
        self.tableWidget_testplan.clearContents()
        self.tableWidget_testplan.setRowCount(len(self.excel_data)-1)
        self.tableWidget_testplan.setColumnCount(len(self.excel_data[0]))

        for i, row in enumerate(self.excel_data):
            if i == 0:
                self.tableWidget_testplan.setHorizontalHeaderLabels(row)
            else:
                for j, col in enumerate(row):
                    if str(col) == "None":
                        col = ""
                    if j == 0 :
                        chkBoxItem = QTableWidgetItem(str(col))
                        chkBoxItem.setText(str(col))
                        chkBoxItem.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                        chkBoxItem.setCheckState(QtCore.Qt.Unchecked)
                        self.tableWidget_testplan.setItem(i-1, j, chkBoxItem)
                    else:    
                        item = QTableWidgetItem(str(col))
                        self.tableWidget_testplan.setItem(i-1, j, item)
    
    def load_excel(self, sheet_name):
        sheet = self.wb[sheet_name]
        self.excel_data = []
        try:
            for row in sheet.values:
                try:
                    self.excel_data.append(row)
                except Exception:
                    continue
            try:
                self.w2table()
                #self.w2measure()
            except Exception as e: 
                print(e)
                logging.error("w2table error!")
        except Exception as e:
            logging.error("Loading Excel sheet Error !" , e)
    
    def read_sheetnames(self, CSV_path):
        self.wb =  load_workbook(CSV_path)
        sheetnames = self.wb.sheetnames
        return sheetnames

    def set_comboBox_sheet(self, combobox, sheet_list):
        combobox.clear()
        combobox.addItems(sheet_list)
    
    def open_testplan(self):
        try:
            filename, filetype = QFileDialog.getOpenFileName(self, "Open file", "./", "Excel (*.xlsx *.xls)")
            if filename != "":
                self.lineEdit_testplan.setText(filename)
                #self.settings.setValue('SETUP/TESTPLAN_PATH', filename) 
                sheetnames = self.read_sheetnames(self.lineEdit_testplan.text())
                self.set_comboBox_sheet(self.comboBox_testplan, sheetnames)
                self.switch_index = 0
                #print(self.test_item[0])
                self.load_excel(self.test_item[0])

        except Exception as e:
            logging.error(e)
            logging.error("Open Excel File Error ! ")
    
    def open_template(self):
        try:
            filename, filetype = QFileDialog.getOpenFileName(self, "Open file", "./", "Word (*.docx)")
            if filename != "":
                self.lineEdit_template.setText(filename)
        except Exception as e:
            print(e)
            logging.error("Open Word File Error ! ")

    def open_uploadfile(self):
        try:
            filename, filetype = QFileDialog.getOpenFileName(self, "Open file", "./", "Excel (*.xlsx *.xls)")
            if filename != "":
                self.lineEdit_upload.setText(filename)

                upload_sys = Upload_Runthread()
                upload_sys.file_path = self.lineEdit_upload.text()
                upload_sys.tableWidget_upload = self.tableWidget_upload
                upload_sys.run()

        except Exception as e:
            print(e)
            logging.error("Open Word File Error ! ")

    def getusblist(self):
        try:
            rm = visa.ResourceManager()
            usb_list = rm.list_resources()
            rm.close
        except Exception:
            usb_list = []
        
        return usb_list
    
    def setup_TK_addres(self):
       self.Refresh_VISA(self.comboBox_TK_visa)
    def setup_PS_addres(self):
        self.Refresh_VISA(self.comboBox_PS_visa)
    def setup_EL_addres(self):
        self.Refresh_VISA(self.comboBox_EL_visa)

    def Refresh_VISA(self, combobox_obj ):
        usb_list = self.getusblist()
        combobox_obj.clear()
        logging.info(usb_list)
        try:
            if len(usb_list) != 0:
                combobox_obj.addItems(usb_list)
            else:
                combobox_obj.addItems([""])
        except Exception as e:
            logging.error(e)
            combobox_obj.addItems([""])
    
    def respones2table(self, msg):
        #print(msg)
        index_dic = {"Line":11,"Regulation": 14 , "Load": 10,  "Eff":6}
        index_shift = index_dic[self.test_item[self.switch_index]]

        for index, col in enumerate(msg[1]):
            item = QTableWidgetItem(str(col))
            #print(msg[0]-1, index+index_shift)
            self.tableWidget_testplan.setItem(msg[0]-1, index+index_shift , item)

    def thread_stop(self):
        logging.info("Cancel Generate Report Progress")
        self.thread.terminate()

if __name__ == "__main__":
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    app = QtWidgets.QApplication(sys.argv)
    AutoRepor = mainProgram()
    AutoRepor.show()
    sys.exit(app.exec_())