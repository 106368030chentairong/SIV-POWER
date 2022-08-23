import os, sys, time
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import logging
from openpyxl import load_workbook, Workbook

class Upload_Runthread():
    _info = pyqtSignal(list)

    def __init__(self):
        super(Upload_Runthread, self).__init__()
        self.file_path = None
        self.tableWidget_upload = None

    def load_excel(self, file_path , sheet_name):
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        excel_data = []
        try:
            for row in sheet.values:
                try:
                    excel_data.append(row)
                except Exception:
                    continue
            try:
                self.w2table()
                #self.w2measure()
            except Exception as e: 
                print(e)
                logging.error("w2table error!")
        except Exception as e:
            logging.error("Loading Excel sheet Error !" , e)
    
    def load_test_data(self, file_path , sheet_name):
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]

        for row in sheet.values:
            print(row)

    def w2table(self):
        self.tableWidget_upload.clearContents()
        self.tableWidget_upload.setRowCount(len(self.excel_data)-1)
        self.tableWidget_upload.setColumnCount(len(self.excel_data[0]))

        for i, row in enumerate(self.excel_data):
            if i == 0:
                self.tableWidget_upload.setHorizontalHeaderLabels(row)
            else:
                for j, col in enumerate(row):
                    if str(col) == "None":
                        col = ""
                    item = QTableWidgetItem(str(col))
                    self.tableWidget_upload.setItem(i-1, j, item)

    def run(self):
        self.load_excel("./EE_TestReport_Format.xlsx", "sample data")
        self.load_test_data(self.file_path, "Testing")